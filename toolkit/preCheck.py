import sys
import configparser
from web3 import Web3
from openpyxl import load_workbook

# Read the config
config = configparser.ConfigParser()
config.read('config.ini')
inputFile = config['data']['inputfile']
blockchainRpcEndpoint = config['blockchain']['rpc']
blockchainChainId = int(config['blockchain']['chainid'])

# Connect to blockchain
w3 = Web3(Web3.HTTPProvider(blockchainRpcEndpoint))
if(w3.isConnected()):
    print(blockchainRpcEndpoint)
    print(blockchainChainId)
else:
    print('Unable to connect to {0}'.format(blockchainRpcEndpoint))
# Open the spreadsheet
workbook = load_workbook(filename = inputFile, data_only=True)
worksheet = workbook.active

totalToTransfer = 0
totalRows = 0
# Pre check
for i in range(1, worksheet.max_row+1):
    recipientAddress_temp = worksheet.cell(row=i, column=2).value
    if isinstance(recipientAddress_temp, str):
        if(recipientAddress_temp.startswith('0x')):
            print("Trying row {0}\n".format(i))
            try:
                recipientAddress = w3.toChecksumAddress(recipientAddress_temp)
                if w3.isAddress(recipientAddress):
                    raw_amount = worksheet.cell(row=i, column=3).value
                    #print(amount_temp)
                    amount = int(float(raw_amount))
                    if amount > 0:
                        print("Correct!\nAddress: {0} \nValue: {1}\n\n".format(recipientAddress, amount))
                        totalToTransfer = totalToTransfer + amount
                        totalRows = totalRows + 1
                    else:
                        print("FAILED!\nAddress: {0} \nValue: {1}\n\n".format(recipientAddress, amount))
                        sys.exit(1)
                else:
                    print("FAILED!\nAddress on row {0} is not valid\n\n".format(i))
                    sys.exit(1)
            except:
                print("Error on row {0}, please check spreadsheet and try again\n\n".format(i))
                sys.exit(1)
        else:
            print("FAILED!\nAddress on row {0} must start with 0x\n\n".format(i))
            sys.exit(1)
    else:
        print("FAILED!\nAddress on row {0} must be string\n\n".format(i))
        sys.exit(1)
print("Total rows processed is {0}".format(totalRows))
print("Total amount to be transferred in wei is {0}".format(totalToTransfer))
print("Total amount to be transferred in whole tokens is {}".format(w3.fromWei(totalToTransfer, 'ether')))
print("\n\nONLY RUN THE BULK DEPOSIT TOKENS SCRIPT IF THESE NUMBERS ARE CORRECT!\n\n")
