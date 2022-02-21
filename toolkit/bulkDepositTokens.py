# Things this program needs
# pip3 install openpyxl
# pip3 install configparser
# pip3 install web3
# How to run this script
# python3 send_tokens.py

# Import config parser and web3 and spreadsheet parser 
import sys
import time
import json
import pickle
import configparser
from web3 import Web3
from datetime import datetime
from openpyxl import load_workbook


abi = '''[
    {
        "inputs": [
            {
                "internalType": "contract IERC20",
                "name": "_erc20_contract_address",
                "type": "address"
            }
        ],
        "stateMutability": "nonpayable",
        "type": "constructor"
    },
    {
        "anonymous": false,
        "inputs": [
            {
                "indexed": false,
                "internalType": "address",
                "name": "recipient",
                "type": "address"
            },
            {
                "indexed": false,
                "internalType": "uint256",
                "name": "amount",
                "type": "uint256"
            }
        ],
        "name": "AllocationPerformed",
        "type": "event"
    },
    {
        "anonymous": false,
        "inputs": [
            {
                "indexed": false,
                "internalType": "address",
                "name": "from",
                "type": "address"
            },
            {
                "indexed": false,
                "internalType": "uint256",
                "name": "amount",
                "type": "uint256"
            }
        ],
        "name": "TokensDeposited",
        "type": "event"
    },
    {
        "anonymous": false,
        "inputs": [
            {
                "indexed": false,
                "internalType": "address",
                "name": "recipient",
                "type": "address"
            },
            {
                "indexed": false,
                "internalType": "uint256",
                "name": "amount",
                "type": "uint256"
            }
        ],
        "name": "TokensUnlocked",
        "type": "event"
    },
    {
        "inputs": [
            {
                "internalType": "uint256",
                "name": "_timePeriod",
                "type": "uint256"
            }
        ],
        "name": "accelerateForTesting",
        "outputs": [],
        "stateMutability": "nonpayable",
        "type": "function"
    },
    {
        "inputs": [],
        "name": "accelerateFunctionDisarmed",
        "outputs": [
            {
                "internalType": "bool",
                "name": "",
                "type": "bool"
            }
        ],
        "stateMutability": "view",
        "type": "function"
    },
    {
        "inputs": [],
        "name": "allIncomingDepositsFinalised",
        "outputs": [
            {
                "internalType": "bool",
                "name": "",
                "type": "bool"
            }
        ],
        "stateMutability": "view",
        "type": "function"
    },
    {
        "inputs": [
            {
                "internalType": "address",
                "name": "",
                "type": "address"
            }
        ],
        "name": "alreadyWithdrawn",
        "outputs": [
            {
                "internalType": "uint256",
                "name": "",
                "type": "uint256"
            }
        ],
        "stateMutability": "view",
        "type": "function"
    },
    {
        "inputs": [
            {
                "internalType": "address",
                "name": "",
                "type": "address"
            }
        ],
        "name": "balances",
        "outputs": [
            {
                "internalType": "uint256",
                "name": "",
                "type": "uint256"
            }
        ],
        "stateMutability": "view",
        "type": "function"
    },
    {
        "inputs": [
            {
                "internalType": "address[]",
                "name": "recipients",
                "type": "address[]"
            },
            {
                "internalType": "uint256[]",
                "name": "amounts",
                "type": "uint256[]"
            }
        ],
        "name": "bulkDepositTokens",
        "outputs": [],
        "stateMutability": "nonpayable",
        "type": "function"
    },
    {
        "inputs": [],
        "name": "contractBalance",
        "outputs": [
            {
                "internalType": "uint256",
                "name": "",
                "type": "uint256"
            }
        ],
        "stateMutability": "view",
        "type": "function"
    },
    {
        "inputs": [
            {
                "internalType": "address",
                "name": "recipient",
                "type": "address"
            },
            {
                "internalType": "uint256",
                "name": "amount",
                "type": "uint256"
            }
        ],
        "name": "depositTokens",
        "outputs": [],
        "stateMutability": "nonpayable",
        "type": "function"
    },
    {
        "inputs": [],
        "name": "disarmAcceleratorFunctionality",
        "outputs": [],
        "stateMutability": "nonpayable",
        "type": "function"
    },
    {
        "inputs": [],
        "name": "erc20Contract",
        "outputs": [
            {
                "internalType": "contract IERC20",
                "name": "",
                "type": "address"
            }
        ],
        "stateMutability": "view",
        "type": "function"
    },
    {
        "inputs": [],
        "name": "finalizeAllIncomingDeposits",
        "outputs": [],
        "stateMutability": "nonpayable",
        "type": "function"
    },
    {
        "inputs": [],
        "name": "initialTimestamp",
        "outputs": [
            {
                "internalType": "uint256",
                "name": "",
                "type": "uint256"
            }
        ],
        "stateMutability": "view",
        "type": "function"
    },
    {
        "inputs": [],
        "name": "owner",
        "outputs": [
            {
                "internalType": "address",
                "name": "",
                "type": "address"
            }
        ],
        "stateMutability": "view",
        "type": "function"
    },
    {
        "inputs": [
            {
                "internalType": "uint256",
                "name": "_timePeriod",
                "type": "uint256"
            }
        ],
        "name": "revertAccelerator",
        "outputs": [],
        "stateMutability": "nonpayable",
        "type": "function"
    },
    {
        "inputs": [
            {
                "internalType": "uint256",
                "name": "_timePeriod",
                "type": "uint256"
            }
        ],
        "name": "setTimestamp",
        "outputs": [],
        "stateMutability": "nonpayable",
        "type": "function"
    },
    {
        "inputs": [],
        "name": "timePeriod",
        "outputs": [
            {
                "internalType": "uint256",
                "name": "",
                "type": "uint256"
            }
        ],
        "stateMutability": "view",
        "type": "function"
    },
    {
        "inputs": [],
        "name": "timestampSet",
        "outputs": [
            {
                "internalType": "bool",
                "name": "",
                "type": "bool"
            }
        ],
        "stateMutability": "view",
        "type": "function"
    },
    {
        "inputs": [
            {
                "internalType": "contract IERC20",
                "name": "token",
                "type": "address"
            },
            {
                "internalType": "address",
                "name": "to",
                "type": "address"
            },
            {
                "internalType": "uint256",
                "name": "amount",
                "type": "uint256"
            }
        ],
        "name": "transferAccidentallyLockedTokens",
        "outputs": [],
        "stateMutability": "nonpayable",
        "type": "function"
    },
    {
        "inputs": [
            {
                "internalType": "contract IERC20",
                "name": "token",
                "type": "address"
            },
            {
                "internalType": "address",
                "name": "to",
                "type": "address"
            },
            {
                "internalType": "uint256",
                "name": "amount",
                "type": "uint256"
            }
        ],
        "name": "transferTimeLockedTokensAfterTimePeriod",
        "outputs": [],
        "stateMutability": "nonpayable",
        "type": "function"
    },
    {
        "inputs": [
            {
                "internalType": "uint256",
                "name": "amount",
                "type": "uint256"
            },
            {
                "internalType": "address payable",
                "name": "ownersAddress",
                "type": "address"
            }
        ],
        "name": "withdrawEth",
        "outputs": [],
        "stateMutability": "nonpayable",
        "type": "function"
    },
    {
        "stateMutability": "payable",
        "type": "receive"
    }
]'''

abiObject = json.loads(abi)

# Read the config
config = configparser.ConfigParser()
config.read('config.ini')
sendersPrivateKey = config['sender']['privatekey']
sendersPublicKey = config['sender']['publickey']
blockchainRpcEndpoint = config['blockchain']['rpc']
blockchainChainId = int(config['blockchain']['chainid'])
inputFile = config['data']['inputfile']
timelockContractAddress = config['contract']['timelockaddress']

# Connect to blockchain
w3 = Web3(Web3.HTTPProvider(blockchainRpcEndpoint))
if(w3.isConnected()):
    print(blockchainRpcEndpoint)
    print(blockchainChainId)
else:
    print('Unable to connect to {0}'.format(blockchainRpcEndpoint))

timelockContractInstance = w3.eth.contract(address=timelockContractAddress, abi=abi)

# Helper functions
def writeData(_file_name, _data_to_write):
    with open(_file_name, 'wb') as f:
        pickle.dump(_data_to_write, f)

def readData(_file_name):
    with open(_file_name, 'rb') as f:
        return pickle.load(f)

# Open the spreadsheet
workbook = load_workbook(filename = inputFile, data_only=True)
worksheet = workbook.active

# Create blank report object
outerReportObject = []

# Create unique file name for this run
timeOfRunning = datetime.today().strftime('%Y-%m-%d-%H-%M-%S')

writeData(timeOfRunning, outerReportObject)

class GasUsage:
    def __init__(self):
        self.cumulativeGasUsed = 0

    def incrementGas(self, _gas_used):
        self.cumulativeGasUsed = self.cumulativeGasUsed + _gas_used

    def reportGasUsed(self):
        return self.cumulativeGasUsed

def confirmTransactions(_list_of_transaction_hashes):
    print("Confirming a set of transactions, please wait ...")
    for i in _list_of_transaction_hashes:
        w3.eth.wait_for_transaction_receipt(i)
    print("Transactions confirmed, continuing now ...")

def executeTransaction(_gasUsage, _sendersPrivateKey, _sendersPublicKey, _blockchainChainId, _nonce, _timelockContractInstance, _tempKeyValueDict):
    transactionObject = {"nonce": _nonce, "from": _sendersPublicKey, "chainId": _blockchainChainId}
    print("Transaction object: {0}".format(json.dumps(transactionObject)))
    print("Creating transaction object")
    addresses = []
    amounts = []
    for k, v in _tempKeyValueDict.items():
        addresses.append(k)
        amounts.append(v)
    transaction = _timelockContractInstance.functions.bulkDepositTokens(addresses, amounts).buildTransaction(transactionObject)
    print("Signing transaction")
    signedTransaction = w3.eth.account.signTransaction(transaction, private_key=_sendersPrivateKey)
    print("Sending transaction" + str(signedTransaction.rawTransaction))
    sentTransaction = w3.eth.sendRawTransaction(signedTransaction.rawTransaction)
    print(sentTransaction.hex())
    transactionReceipt = w3.eth.wait_for_transaction_receipt(sentTransaction.hex(), timeout=120, poll_latency=0.1)
    print(transactionReceipt)
    transactionReceiptAsJson = json.loads(w3.toJSON(transactionReceipt))
    print("Cumulative Gas used:")
    print(transactionReceiptAsJson["cumulativeGasUsed"])
    print(w3.fromWei(transactionReceiptAsJson["cumulativeGasUsed"], 'ether'))
    _gasUsage.incrementGas(transactionReceiptAsJson["cumulativeGasUsed"])
    return sentTransaction.hex()


# Optional command line argument
gasUsage = GasUsage()
startFrom = False
n = len(sys.argv)
if n == 2:
    startFromAddress = sys.argv[1]
    if(startFromAddress.startswith('0x')):
        if w3.isAddress(startFromAddress):
            startFrom = True
            print("Starting from {0}".format(startFromAddress))
else:
    print("Starting from the beginning of xlsx file")

# Iterate through the spreadsheet
depupedDict = {}
for i in range(1, worksheet.max_row+1):
    recipientAddress_temp = worksheet.cell(row=i, column=2).value
    if isinstance(recipientAddress_temp, str):
        if(recipientAddress_temp.startswith('0x')):
            try:
                print("1");
                recipientAddress = w3.toChecksumAddress(recipientAddress_temp)
                print("2");
                if w3.isAddress(recipientAddress):
                    print("3");
                    print("Processing address: {0}".format(recipientAddress))
                    print("Processing amount: {0}".format(worksheet.cell(row=i, column=3).value))
                    raw_amount = worksheet.cell(row=i, column=3).value
                    print("4");
                    print(raw_amount)
                    amount = int(float(raw_amount))
                    print("5");
                    print("Processing amount: {0}".format(amount))
                    if recipientAddress in depupedDict:
                        print("6");
                        print("We alread have that address: {0}".format(recipientAddress))
                        print("Adding {0} and {1}".format(depupedDict[recipientAddress], amount))
                        temp_amount = depupedDict[recipientAddress] + amount
                        print("7");
                        depupedDict[recipientAddress] = temp_amount
                        print("8");
                    else:
                        depupedDict[recipientAddress] = amount
            except:
                print('Invalid address: {0}'.format(recipientAddress))
print("***** Deduped list of recipient accounts: {0} ******".format(len(depupedDict)))
next_nonce_should_be = w3.eth.get_transaction_count(sendersPublicKey,"pending") + 1
processingCommenced = False
# Iterate through the dict
tempListOfTransactionHashes = []
tempKeyValueDict = {}
nonce = w3.eth.get_transaction_count(sendersPublicKey,"pending")
for key, value in depupedDict.items():
    if gasUsage.reportGasUsed() < 50000000000000000:
        print('startFrom: {0}, processingCommenced{1}'.format(startFrom, processingCommenced))
        if startFrom == True and processingCommenced == False:
            print('{0}{0}'.format(key, startFromAddress))
            if key == startFromAddress:
                processingCommenced = True
            else:
                continue
        else:
            processingCommenced = True
        print("Processing: {0} {1}".format(key, value))
        if value > 0:
            tempKeyValueDict[key] = value
            if len(tempListOfTransactionHashes) == 1:
                confirmTransactions(tempListOfTransactionHashes)
                tempListOfTransactionHashes = []
            ## Amount of addresses and amounts to pass into bulk function
            if len(tempKeyValueDict) == 50:
                transactionHash = executeTransaction(gasUsage, sendersPrivateKey, sendersPublicKey, blockchainChainId, nonce, timelockContractInstance, tempKeyValueDict)
                print("Total cumulative gas used: {0}".format(gasUsage.reportGasUsed()))
                tempListOfTransactionHashes.append(transactionHash)
                nonce = w3.eth.get_transaction_count(sendersPublicKey,"pending")
                print('Nonce: {0}'.format(nonce))
                if nonce < next_nonce_should_be:
                    print("Waiting for nonce value to catchup")
                    time.sleep(15)
                    nonce = w3.eth.get_transaction_count(sendersPublicKey,"pending")
                print("Using nonce: {0}".format(nonce))
                next_nonce_should_be = nonce + 1
                tempOuterReportObject = readData(timeOfRunning)
                for tk, tv in tempKeyValueDict.items():
                    tempInnerReportObject = [tk, tv, transactionHash]
                    tempOuterReportObject.append(tempInnerReportObject)
                writeData(timeOfRunning, tempOuterReportObject)
                tempKeyValueDict = {}
                time.sleep(0.25)
        else:
            print("Amount must be more than zero, please check all amounts and try again")
            continue
    else:
        print("Exceeded the approx amount of gas we wanted to spend, please check gas consumption and try again")
        break

if len(tempKeyValueDict) > 0:
    transactionHash = executeTransaction(gasUsage, sendersPrivateKey, sendersPublicKey, blockchainChainId, nonce, timelockContractInstance, tempKeyValueDict)
    print("Total cumulative gas used: {0}".format(gasUsage.reportGasUsed()))
    tempOuterReportObject = readData(timeOfRunning)
    for tk, tv in tempKeyValueDict.items():
        tempInnerReportObject = [tk, tv, transactionHash]
        tempOuterReportObject.append(tempInnerReportObject)
    writeData(timeOfRunning, tempOuterReportObject)

if len(tempListOfTransactionHashes) > 0:
    confirmTransactions(tempListOfTransactionHashes)
print("Total cumulative gas used: {0}".format(gasUsage.reportGasUsed()))
